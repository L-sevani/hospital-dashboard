"""
procesar_excel.py
-----------------
Lee el Excel de reporte de infecciones (uno o varios periodos) y genera:
  1. base_microorganismos_expandida.xlsx  → una fila por microorganismo
  2. dashboard_data.json                  → JSON listo para pegar en el dashboard HTML

Uso:
    python procesar_excel.py archivo.xlsx
    python procesar_excel.py archivo.xlsx -o salida/
"""

import sys, os, json, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CATÁLOGO: Tipo de infección → Subgrupo del dashboard ─────────────────────
INFECCION_MAP = {
    # NEUMONÍA
    "NEUMONIA ASOCIADA A VENTILADOR": "NEUMONÍA",
    "NEUMONIA CLINICA NO ASOCIADA A VENTILADOR": "NEUMONÍA",

    # HERIDAS
    "INFECCION DE HERIDA QUIRURGICA INCISIONAL SUPERFICIAL": "HERIDAS",
    "INFECCION DE HERIDA QUIRURGICA INCISIONAL PROFUNDA": "HERIDAS",
    "INFECCION DE HERIDA QUIRURGICA DE ORGANOS Y ESPACIOS": "HERIDAS",

    # ITS
    "INFECCIONES DE SITIO DE INSERCION DE CATETER, TUNEL O PUERTO SUBCUTANEO": "ITS",
    "ITS RELACIONADA A CATETER VENOSO CENTRAL": "ITS",
    "ITS CONFIRMADA POR LABORATORIO": "ITS",
    "BACTERIEMIA PRIMARIA NO DEMOSTRADA": "ITS",
    "ITS SECUNDARIA A PROCEDIMIENTOS (CISTOSCOPIAS Y COLANGIOGRAFIAS)": "ITS",
    "ENDOCARDITIS": "ITS",
    "ITS RELACIONADA A CONTAMINACION DE SOLUCIONES": "ITS",
    "INFUSIONES O MEDICAMENTOS INTRAVENOSOS": "ITS",
    "ITS RELACIONADA A CONTAMINACION DE SOLUCIONES, INFUSIONES O MEDICAMENTOS INTRAVENOSOS": "ITS",

    # IVU
    "IVU ASOCIADA A SONDA VESICAL": "IVU",
    "IVU NO ASOCIADA A SONDA VESICAL": "IVU",

    # PERITONITIS
    "PERITONITIS ASOCIADA A DIALISIS": "PERITONITIS",
    "PERITONITIS ASOCIADA A LA INSTALACION DE CATETER DE DIALISIS PERITONEAL": "PERITONITIS",

    # RESTO DE IAAS — todo lo demás
    "INFECCIONES DE PIEL Y TEJIDOS BLANDOS": "RESTO DE IAAS",
    "INFECCIONES DE PIEL Y TEJIDOS BLANDOS EN PACIENTES CON QUEMADURAS": "RESTO DE IAAS",
    "FLEBITIS": "RESTO DE IAAS",
    "CONJUNTIVITIS": "RESTO DE IAAS",
    "OTRO": "RESTO DE IAAS",
    "RINOFARINGITIS Y FARINGOAMIGDALITIS (IVRA)": "RESTO DE IAAS",
    "COVID-19": "RESTO DE IAAS",
    "ENDOMETRITIS": "RESTO DE IAAS",
    "BRONQUITIS": "RESTO DE IAAS",
    "TRAQUEOBRONQUITIS": "RESTO DE IAAS",
    "TRAQUEITIS SIN DATOS DE NEUMONIA": "RESTO DE IAAS",
    "INFECCIONES DE LA BURSA O ARTICULARES": "RESTO DE IAAS",
    "OTITIS MEDIA AGUDA": "RESTO DE IAAS",
    "INFECCIONES RELACIONADAS A PROCEDIMIENTOS ODONTOLOGICOS": "RESTO DE IAAS",
    "FASCITIS NECROSANTE": "RESTO DE IAAS",
    "GANGRENA INFECCIOSA": "RESTO DE IAAS",
    "CELULITIS": "RESTO DE IAAS",
    "MIOSITIS Y LINFADENITIS": "RESTO DE IAAS",
    "GASTROENTERITIS NOSOCOMIAL": "RESTO DE IAAS",
    "MENINGITIS O VENTRICULITIS": "RESTO DE IAAS",
    "INFLUENZA VIRUS": "RESTO DE IAAS",
}

# ─── CATÁLOGO: Tipo de cultivo → Subgrupo del dashboard ───────────────────────
CULTIVO_MAP = {
    "HEMOCULTIVO DURANTE PICO FEBRIL": "HEMOCULTIVOS",
    "SOLO HEMOCULTIVO CENTRAL": "HEMOCULTIVOS",
    "CULTIVO DE PUNTA DE CATETER + PERIFERICO": "HEMOCULTIVOS",
    "HEMOCULTIVO SIN PICO FEBRIL": "HEMOCULTIVOS",
    "SOLO HEMOCULTIVO PERIFERICO": "HEMOCULTIVOS",
    "HEMOCULTIVO CENTRAL + PERIFERICO": "HEMOCULTIVOS",
    "HEMOCULTIVO PERIFERICO (1 BRAZO) - INADECUADO": "HEMOCULTIVOS",
    "HEMOCULTIVOS PERIFERICOS (2 BRAZOS)": "HEMOCULTIVOS",
    "HEMOCULTIVOS": "HEMOCULTIVOS",
    "SOLO CULTIVO DE PUNTA DE CATETER. - INADECUADO": "HEMOCULTIVOS",
    "UROCULTIVO POR PUNCION DE SONDA": "UROCULTIVOS",
    "UROCULTIVO POR CHORRO MEDIO": "UROCULTIVOS",
    "CULTIVO DE LA SONDA URETRAL - INADECUADO": "UROCULTIVOS",
    "UROCULTIVO POR PUNCION SUPRAPUBICA": "UROCULTIVOS",
    "CULTIVO DE ESPUTO": "NEUMO",
    "CULTIVO DE LIQUIDO PLEUREAL": "NEUMO",
    "ASPIRADO TRAQUEAL": "NEUMO",
    "CULTIVO POR PUNCION ASPIRACION": "NEUMO",
    "LAVADO BRONQUIAL": "NEUMO",
    "CULTIVO DE LIQUIDO PERITONEAL": "PERITONEAL",
    "CULTIVO POR HISOPO - INADECUADO": "OTROS",
    "CULTIVO DEL PUS - INADECUADO": "OTROS",
    "COPROCULTIVO": "OTROS",
    "PONER TODAS LAS OPCIONES": "OTROS",
    "CULTIVO DE DRENAJES - INADECUADO": "OTROS",
    "CULTIVO DE LIQUIDO ESTERIL": "OTROS",
    "EXUDADO NASOFARINGEO": "OTROS",
    "CULTIVOS TOMADOS DE LA BOLSA RECOLECTORA DE ORINA - INADECUADO": "OTROS",
    "CULTIVO DE LCR": "OTROS",
    "OTROS - INADECUADO": "OTROS",
    "CULTIVO DE BIOPSIA": "OTROS",
}

# ─── MAPEO SERVICIOS → SUBGRUPO (igual que el dashboard) ──────────────────────
SERVICIO_MAP = {
    "GINECOLOGIA": "GINECOLOGÍA", "OBSTETRICIA": "GINECOLOGÍA",
    "GINECOLOGIA Y OBSTETRICIA": "GINECOLOGÍA",
    "ONCOLOGIA GINECOLOGICA": "GINECOLOGÍA",
    "ONCOLOGIA DE TUMORES DE MAMA": "GINECOLOGÍA",
    "MEDICINA INTERNA": "MEDICINA INTERNA", "NEFROLOGIA": "MEDICINA INTERNA",
    "HEMODIALISIS": "MEDICINA INTERNA", "MEDICINA FAMILIAR": "MEDICINA INTERNA",
    "CARDIOLOGIA": "MEDICINA INTERNA", "GERONTOLOGIA/GERIATRIA": "MEDICINA INTERNA",
    "NEUROLOGIA": "MEDICINA INTERNA", "AUDIOLOGIA": "MEDICINA INTERNA",
    "REUMATOLOGIA": "MEDICINA INTERNA", "NEUMOLOGIA": "MEDICINA INTERNA",
    "NEFROLOGIA PEDIATRICA": "MEDICINA INTERNA", "PSIQUIATRIA CLINICA": "MEDICINA INTERNA",
    "MEDICINA INTERNA PEDIATRICA": "MEDICINA INTERNA",
    "CIRUGIA GENERAL": "CIRUGÍA", "ORTOPEDIA Y TRAUMATOLOGIA": "CIRUGÍA",
    "TRAUMATOLOGIA": "CIRUGÍA", "UROLOGIA": "CIRUGÍA", "OFTALMOLOGIA": "CIRUGÍA",
    "ORTOPEDIA": "CIRUGÍA", "CIRUGIA PEDIATRICA": "CIRUGÍA",
    "OTORRINOLARINGOLOGIA": "CIRUGÍA", "CIRUGIA AMBULATORIA": "CIRUGÍA",
    "ANGIOLOGIA": "CIRUGÍA", "CIRUGIA PLASTICA RECONSTRUCTIVA": "CIRUGÍA",
    "TRUMATOLOGIA Y ORTOPEDIA DE CADERA Y PELVIS": "CIRUGÍA",
    "ONCOLOGIA QUIRURGICA": "CIRUGÍA", "CIRUGIA MAXILOFACIAL": "CIRUGÍA",
    "CIRUGIA CARDIOVASCULAR": "CIRUGÍA",
    "TRAUMATOLOGIA Y O. DE CIRUGIA TORACO ABDOMINAL": "CIRUGÍA",
    "TRAUMATOLOGIA Y ORTOPEDIA DE MIEMBRO PELVICO": "CIRUGÍA",
    "TRAUMATOLOGIA Y ORTOPEDIA DE MIEMBRO TORACICO": "CIRUGÍA",
    "NEUROCIRUGIA": "CIRUGÍA",
    "CUNERO PATOLOGICO": "PEDIATRÍA", "PEDIATRIA MEDICA": "PEDIATRÍA",
    "MED. ENFERMO EDO. CRITICO": "UCI",
    "URGENCIAS": "URGENCIAS",
}

def normalizar(texto):
    if pd.isna(texto):
        return ""
    return str(texto).strip().upper()

def mapear_infeccion(texto):
    t = normalizar(texto)
    if t in INFECCION_MAP:
        return INFECCION_MAP[t], t
    # Búsqueda parcial por palabras clave
    if t.startswith("ITS"):
        return "ITS", t
    if t.startswith("IVU"):
        return "IVU", t
    if "NEUMONIA" in t or "NEUMONÍA" in t:
        return "NEUMONÍA", t
    if "HERIDA QUIRURGICA" in t or "HERIDAS" in t:
        return "HERIDAS", t
    if "PERITONITIS" in t:
        return "PERITONITIS", t
    return "RESTO DE IAAS", t

def mapear_cultivo(texto):
    t = normalizar(texto)
    if t in CULTIVO_MAP:
        return CULTIVO_MAP[t]
    if "HEMOCULTIVO" in t or "HEMOCULTIVOS" in t:
        return "HEMOCULTIVOS"
    if "UROCULTIVO" in t:
        return "UROCULTIVOS"
    if "ESPUTO" in t or "BRONQUIAL" in t or "TRAQUEAL" in t or "PLEUREAL" in t:
        return "NEUMO"
    if "PERITONEAL" in t:
        return "PERITONEAL"
    return "OTROS"

def mapear_servicio(texto):
    t = normalizar(texto)
    if t in SERVICIO_MAP:
        return SERVICIO_MAP[t]
    for k, v in SERVICIO_MAP.items():
        if k in t:
            return v
    return t  # devuelve tal cual si no se encuentra

def extraer_periodo(fecha):
    """Convierte fecha a string YYYY-MM para agrupar por mes."""
    if pd.isna(fecha):
        return None
    try:
        if isinstance(fecha, str):
            # Intenta formatos dd-mm-yyyy o yyyy-mm-dd
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    from datetime import datetime
                    dt = datetime.strptime(fecha.strip(), fmt)
                    return dt.strftime("%Y-%m")
                except:
                    continue
        import datetime as dt
        if hasattr(fecha, 'strftime'):
            return fecha.strftime("%Y-%m")
    except:
        pass
    return None

def leer_excel(path):
    """Lee el Excel con el formato de IMSS (encabezado real en fila 4, índice 3)."""
    raw = pd.read_excel(path, header=3)
    real_cols = raw.iloc[0].tolist()
    raw.columns = real_cols
    raw = raw.iloc[1:].reset_index(drop=True)
    return raw, real_cols

def expandir_microorganismos(raw, real_cols):
    """Una fila por microorganismo. Ignora filas sin microorganismo."""
    blocks = [
        (41, 42, 43, 44, 112, 113),   # micro 1
        (114, 115, 116, 117, 185, 186), # micro 2
        (187, 188, 189, 190, 258, 259), # micro 3
        (260, 261, 262, 263, 331, 332), # micro 4
    ]
    abx_names = [real_cols[i] for i in range(44, 113)]
    base_cols_idx = list(range(41))

    records = []
    for _, row in raw.iterrows():
        base = {real_cols[i]: row.iloc[i] for i in base_cols_idx}

        # Periodo desde fecha de detección
        fecha = row.iloc[16]  # Fecha de detección de la infección
        if pd.isna(fecha) or extraer_periodo(fecha) is None:
            fecha = row.iloc[0]  # Fecha de registro como fallback
        base['_periodo'] = extraer_periodo(fecha)

        # Subgrupos normalizados
        tipo_inf = row.iloc[34]
        sg_inf, tipo_inf_norm = mapear_infeccion(tipo_inf)
        base['_subgrupo_infeccion'] = sg_inf
        base['_tipo_infeccion_norm'] = tipo_inf_norm

        tipo_cult = row.iloc[38]
        base['_subgrupo_cultivo'] = mapear_cultivo(tipo_cult)

        servicio = row.iloc[27]
        base['_subgrupo_servicio'] = mapear_servicio(servicio)

        for num, (res_i, clave_i, micro_i, abx_s, abx_e, mec_i) in enumerate(blocks, 1):
            micro = row.iloc[micro_i]
            if pd.isna(micro) or str(micro).strip() in ('', '0', 'nan'):
                continue
            rec = dict(base)
            rec['N_Micro'] = num
            rec['Resistencia'] = row.iloc[res_i]
            rec['Clave_Grupo'] = row.iloc[clave_i]
            rec['Microorganismo'] = str(micro).strip().upper()
            for j, abx in enumerate(abx_names):
                val = row.iloc[abx_s + j]
                sv = str(val).strip().upper() if pd.notna(val) else ''
                rec[abx] = sv if sv in ('S', 'R', 'I') else ''
            rec['Mecanismo_Resistencia'] = row.iloc[mec_i]
            records.append(rec)

    return pd.DataFrame(records), abx_names

def guardar_excel(df, abx_names, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'BASE EXPANDIDA'

    keep_base = [
        'Fecha de Registro', 'Delegación', 'Unidad', 'NSS',
        'Apellido Paterno', 'Apellido Materno', 'Nombre(s)', 'Edad', 'Sexo',
        'Servicio de cama censable', '_subgrupo_servicio',
        'Fecha de detección de la infección',
        'Tipo de infección', '_subgrupo_infeccion',
        'Tipo de cultivo', '_subgrupo_cultivo',
        'Clasificación de la infección', '_periodo',
        'N_Micro', 'Resistencia', 'Clave_Grupo', 'Microorganismo',
    ]
    useful_abx = [a for a in abx_names if a in df.columns and
                  df[a].isin(['S','R','I']).any()]
    cols = [c for c in keep_base if c in df.columns] + useful_abx + ['Mecanismo_Resistencia']
    df_out = df[[c for c in cols if c in df.columns]].copy()

    hdr_fill  = PatternFill('solid', start_color='1F3864')
    meta_fill = PatternFill('solid', start_color='2C5282')
    w_font    = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    d_font    = Font(name='Arial', size=9)
    b_font    = Font(bold=True, name='Arial', size=9)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Border(*[Side(style='thin')]*4)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    s_fill = PatternFill('solid', start_color='C6EFCE')
    r_fill = PatternFill('solid', start_color='FFC7CE')
    i_fill = PatternFill('solid', start_color='FFEB9C')
    m_fill = PatternFill('solid', start_color='DBEAFE')

    # Header
    for ci, col in enumerate(df_out.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = w_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin

    # Data
    for ri, (_, row) in enumerate(df_out.iterrows(), 2):
        alt = ri % 2 == 0
        row_fill = PatternFill('solid', start_color='EEF2FF' if alt else 'FFFFFF')
        for ci, col in enumerate(df_out.columns, 1):
            val = row[col]
            val = '' if pd.isna(val) else val
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = b_font if col == 'Microorganismo' else d_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin
            if col in useful_abx:
                sv = str(val).strip().upper()
                cell.fill = s_fill if sv=='S' else r_fill if sv=='R' else i_fill if sv=='I' else row_fill
            elif col == 'Microorganismo':
                cell.fill = m_fill
            else:
                cell.fill = row_fill

    # Widths
    for ci, col in enumerate(df_out.columns, 1):
        ltr = get_column_letter(ci)
        if col in useful_abx:
            ws.column_dimensions[ltr].width = 14
        elif col in ('Microorganismo', 'Nombre(s)', 'Tipo de infección'):
            ws.column_dimensions[ltr].width = 24
        elif col in ('Servicio de cama censable', '_subgrupo_servicio', '_subgrupo_infeccion'):
            ws.column_dimensions[ltr].width = 20
        else:
            ws.column_dimensions[ltr].width = 16

    ws.row_dimensions[1].height = 45
    ws.freeze_panes = 'A2'

    # Leyenda
    ws2 = wb.create_sheet('LEYENDA')
    ws2['A1'] = 'LEYENDA'
    ws2['A1'].font = Font(bold=True, name='Arial', size=12)
    for i, (code, desc, color) in enumerate([
        ('S', 'Sensible', 'C6EFCE'),
        ('R', 'Resistente', 'FFC7CE'),
        ('I', 'Intermedio', 'FFEB9C'),
        ('(vacío)', 'No aplica / No realizado', 'FFFFFF'),
    ], 3):
        ws2.cell(row=i, column=1, value=code).fill = PatternFill('solid', start_color=color)
        ws2.cell(row=i, column=2, value=desc)
        ws2.cell(row=i, column=1).font = Font(bold=True, name='Arial', size=10)
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 40

    wb.save(out_path)
    print(f"✅ Excel guardado: {out_path} ({len(df_out)} registros, {len(useful_abx)} antibióticos)")

HOSPITALES_ORDEN = [
    "HGZ 1 TLAXCALA",
    "HGZMF 2 APIZACO",
    "HGSMF 8 TLAXCALA",
]

def construir_json_dashboard(df, abx_names):
    """
    Construye el objeto RAW compatible con el dashboard HTML.
    Agrupa por hospital (Unidad), luego por periodo.
    """
    useful_abx = [a for a in abx_names if a in df.columns and df[a].isin(['S','R','I']).any()]

    # Columna Unidad normalizada
    col_unidad = None
    for c in df.columns:
        if str(c).strip().upper() == 'UNIDAD':
            col_unidad = c
            break
    if col_unidad is None:
        raise ValueError("No se encontró la columna 'Unidad' en el DataFrame.")

    df = df.copy()
    df['_unidad'] = df[col_unidad].astype(str).str.strip().str.upper()

    hospitales_presentes = [h for h in HOSPITALES_ORDEN if h in df['_unidad'].unique()]
    # Agregar cualquier hospital no listado explícitamente
    for h in df['_unidad'].unique():
        if h not in hospitales_presentes and h not in ('', 'NAN'):
            hospitales_presentes.append(h)

    periodos = sorted(df['_periodo'].dropna().unique())

    def sri_por_micro(sub):
        """Para cada microorganismo en sub, calcula totales S/R/I por antibiótico."""
        result = {}
        for micro, grp in sub.groupby('Microorganismo'):
            abx_data = []
            for abx in useful_abx:
                if abx not in grp.columns:
                    continue
                vals = grp[abx].dropna().str.strip().str.upper()
                s = int((vals == 'S').sum())
                r = int((vals == 'R').sum())
                i = int((vals == 'I').sum())
                t = s + r + i
                if t > 0:
                    abx_data.append({"a": abx, "S": s, "R": r, "I": i, "t": t})
            if abx_data:
                result[micro] = {"total": len(grp), "antibioticos": abx_data}
        return result

    def build_section(sub):
        """Construye data + subgrupos para una sección (infecciones o cultivos)."""
        return sri_por_micro(sub)

    def build_periodos(h_df):
        """Construye data_periodos para un hospital."""
        h_periodos = sorted(h_df['_periodo'].dropna().unique())
        data_periodos = {}
        for periodo in h_periodos:
            p_df = h_df[h_df['_periodo'] == periodo]

            # INFECCIONES
            inf_sgs = ['HERIDAS', 'ITS', 'IVU', 'NEUMONÍA', 'PERITONITIS', 'RESTO DE IAAS']
            inf_data = {}
            for sg in inf_sgs:
                sub = p_df[p_df['_subgrupo_infeccion'] == sg]
                if len(sub) > 0:
                    inf_data[sg] = build_section(sub)

            # CULTIVOS
            cult_sgs = ['HEMOCULTIVOS', 'UROCULTIVOS', 'NEUMO', 'PERITONEAL', 'OTROS']
            cult_data = {}
            for sg in cult_sgs:
                sub = p_df[p_df['_subgrupo_cultivo'] == sg]
                if len(sub) > 0:
                    cult_data[sg] = build_section(sub)

            # SERVICIOS
            serv_sgs = ['GINECOLOGÍA', 'MEDICINA INTERNA', 'CIRUGÍA', 'PEDIATRÍA', 'UCI', 'URGENCIAS']
            serv_data = {}
            for sg in serv_sgs:
                sub = p_df[p_df['_subgrupo_servicio'] == sg]
                if len(sub) > 0:
                    serv_data[sg] = build_section(sub)

            data_periodos[periodo] = {
                "infecciones": {"subgrupos": [sg for sg in inf_sgs if sg in inf_data], "data": inf_data},
                "cultivos":    {"subgrupos": [sg for sg in cult_sgs if sg in cult_data], "data": cult_data},
                "servicios":   {"subgrupos": [sg for sg in serv_sgs if sg in serv_data], "data": serv_data},
            }
        return {"periodos": h_periodos, "data": data_periodos}

    hospitales_data = {}
    for h in hospitales_presentes:
        h_df = df[df['_unidad'] == h]
        hospitales_data[h] = build_periodos(h_df)
        print(f"   🏥 {h}: {len(h_df)} microorganismos, periodos: {hospitales_data[h]['periodos']}")

    return {
        "hospitales": hospitales_presentes,
        "periodos": periodos,            # todos los periodos combinados
        "data": hospitales_data,         # data[hospital][periodos/data]
    }

def main():
    if len(sys.argv) < 2:
        print("Uso: python procesar_excel.py archivo.xlsx [-o carpeta_salida]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    out_dir = "."
    if "-o" in sys.argv:
        idx = sys.argv.index("-o")
        out_dir = sys.argv[idx + 1]
        os.makedirs(out_dir, exist_ok=True)

    print(f"📂 Leyendo: {xlsx_path}")
    raw, real_cols = leer_excel(xlsx_path)
    print(f"   Filas encontradas: {len(raw)}")

    df, abx_names = expandir_microorganismos(raw, real_cols)
    print(f"   Registros con microorganismo: {len(df)}")
    if len(df) == 0:
        print("⚠️  No se encontraron filas con microorganismos. Revisa el archivo.")
        sys.exit(1)

    periodos = df['_periodo'].dropna().unique()
    print(f"   Periodos detectados: {sorted(periodos)}")

    # Diagnóstico de infecciones no mapeadas
    no_mapeadas = df[df['_subgrupo_infeccion'] == 'RESTO DE IAAS']['_tipo_infeccion_norm'].unique()
    if len(no_mapeadas):
        print(f"   ⚠️  Infecciones clasificadas como RESTO DE IAAS: {list(no_mapeadas[:5])}")

    excel_out = os.path.join(out_dir, "base_microorganismos_expandida.xlsx")
    guardar_excel(df, abx_names, excel_out)

    json_out = os.path.join(out_dir, "dashboard_data.json")
    data_json = construir_json_dashboard(df, abx_names)
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(data_json, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON guardado: {json_out}")
    print(f"\n📊 Resumen:")
    for h, hdata in data_json["data"].items():
        for p in hdata["periodos"]:
            h_df = df[(df['Unidad'].astype(str).str.strip().str.upper() == h) & (df['_periodo'] == p)]
            print(f"   [{h}] {p}: {len(h_df)} microorganismos")

if __name__ == "__main__":
    main()
